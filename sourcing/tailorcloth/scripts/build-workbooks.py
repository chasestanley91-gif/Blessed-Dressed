import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, ".data")   # produced by extract.mjs
OUTDIR = os.path.dirname(HERE)
os.makedirs(OUTDIR, exist_ok=True)

NAVY = "1F3352"
LIGHT = "EAEFF5"
YELLOW = "FFFF00"
GREY = "F5F5F5"

FONT = "Arial"
H1 = Font(name=FONT, size=14, bold=True, color="FFFFFF")
HDR = Font(name=FONT, size=10, bold=True, color="FFFFFF")
BOLD = Font(name=FONT, size=10, bold=True)
BODY = Font(name=FONT, size=10)
SMALL = Font(name=FONT, size=9, italic=True, color="555555")
TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
BAND_FILL = PatternFill("solid", fgColor=LIGHT)
INPUT_FILL = PatternFill("solid", fgColor=YELLOW)
SEC_FILL = PatternFill("solid", fgColor=GREY)
THIN = Side(style="thin", color="B7C0CC")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def banner(ws, text, ncols, sub=None):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(row=1, column=1, value=text)
    c.font, c.fill = H1, TITLE_FILL
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 26
    if sub:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        s = ws.cell(row=2, column=2 - 1, value=sub)
        s.font = SMALL
        s.alignment = Alignment(vertical="center", indent=1)
        ws.row_dimensions[2].height = 16


def header_row(ws, row, headers, widths):
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font, c.fill, c.border = HDR, HDR_FILL, BOX
        c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 22
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK 1 — CRAFT OPTIONS CATALOGUE
# ══════════════════════════════════════════════════════════════════════════
products = json.load(open(os.path.join(DATA, "craft-options.json")))

wb = Workbook()
ws = wb.active
ws.title = "Overview"
banner(ws, "BLESSED & DRESSED  —  CRAFT OPTIONS CATALOGUE", 6)
ws.cell(row=2, column=1, value="Complete construction and make-up option set. One tab per garment type. Codes are our internal specification codes and appear on every order form.").font = SMALL
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

header_row(ws, 4, ["Garment", "Tab Name", "Option Groups", "Total Options",
                   "Options With Tech Illustration", "Reference Retail (USD)"],
           [26, 20, 16, 15, 30, 22])

r = 5
for p in products:
    ill = sum(1 for x in p["rows"] if x["illustration"] == "Yes")
    vals = [p["name"], p["name"][:31], p["sections"], len(p["rows"]), ill, p["basePrice"]]
    for i, v in enumerate(vals, start=1):
        c = ws.cell(row=r, column=i, value=v)
        c.font, c.border = BODY, BOX
        if r % 2 == 0:
            c.fill = BAND_FILL
        if i == 6:
            c.number_format = '$#,##0.00'
        if i in (3, 4, 5):
            c.alignment = Alignment(horizontal="center")
    r += 1

tot = ws.cell(row=r, column=1, value="TOTAL")
tot.font = BOLD
tot.border = BOX
for i, col in ((3, "C"), (4, "D"), (5, "E")):
    c = ws.cell(row=r, column=i, value=f"=SUM({col}5:{col}{r-1})")
    c.font, c.border = BOLD, BOX
    c.alignment = Alignment(horizontal="center")
ws.cell(row=r, column=2).border = BOX
ws.cell(row=r, column=6).border = BOX

note = ws.cell(row=r + 2, column=1,
               value="Reference Retail is our current U.S. direct-to-consumer entry price for the garment type, shown for context only — it is not a target cost.")
note.font = SMALL
ws.merge_cells(start_row=r + 2, start_column=1, end_row=r + 2, end_column=6)
ws.cell(row=r + 3, column=1,
        value="Every option marked 'Yes' under Tech Illustration has an existing dimensioned construction drawing on file, available on request.").font = SMALL
ws.merge_cells(start_row=r + 3, start_column=1, end_row=r + 3, end_column=6)

HEADERS = ["Section", "Option Category", "Specification Code", "Option Name",
           "Variant Group", "House Default", "Tech Illustration", "Photo Reference",
           "Description / Construction Note"]
WIDTHS = [26, 26, 30, 30, 16, 13, 15, 15, 95]

for p in products:
    s = wb.create_sheet(p["name"][:31])
    banner(s, f"CRAFT OPTIONS  —  {p['name'].upper()}", len(HEADERS))
    s.cell(row=2, column=1,
           value=f"{len(p['rows'])} options across {p['sections']} option groups. 'Specification Code' is the value transmitted on every Blessed & Dressed order.").font = SMALL
    s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(HEADERS))
    header_row(s, 4, HEADERS, WIDTHS)

    rr = 5
    last_section = None
    for row in p["rows"]:
        vals = [row["section"], row["field"], row["code"], row["name"], row["group"],
                row["isDefault"], row["illustration"], row["photo"], row["description"]]
        band = row["section"] != last_section
        for i, v in enumerate(vals, start=1):
            c = s.cell(row=rr, column=i, value=v)
            c.font, c.border = BODY, BOX
            c.alignment = Alignment(vertical="top", wrap_text=(i == 9), indent=1)
            if i in (6, 7, 8):
                c.alignment = Alignment(horizontal="center", vertical="top")
            if i <= 2:
                c.fill = BAND_FILL
        last_section = row["section"]
        rr += 1
    s.auto_filter.ref = f"A4:{get_column_letter(len(HEADERS))}{rr-1}"

# openpyxl writes formulas without cached results, so tell the reader to compute
# them on open rather than showing blanks.
wb.calculation.fullCalcOnLoad = True
wb.save(os.path.join(OUTDIR, "Blessed-and-Dressed_Craft-Options-Catalogue.xlsx"))
print("craft options workbook written")


# ══════════════════════════════════════════════════════════════════════════
# WORKBOOK 2 — SAMPLE ORDER FORM
# ══════════════════════════════════════════════════════════════════════════
meas = json.load(open(os.path.join(DATA, "measurements.json")))

wb2 = Workbook()

# ---- Sheet: How To Use ----
ws = wb2.active
ws.title = "How To Use"
banner(ws, "BLESSED & DRESSED  —  SAMPLE ORDER FORM", 3)
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 100
ws.column_dimensions["C"].width = 20

lines = [
    ("Purpose", "This is a filled example of the order packet Blessed & Dressed transmits per garment. Every order is cut to one customer — there is no size-run."),
    ("Example order", "The values throughout this workbook are a realistic worked example (Order BD-2026-0417). They show the format, not a live order."),
    ("Yellow cells", "Yellow-filled cells are the fields completed per order. Everything else is fixed template text."),
    ("Tab: Order Header", "Customer, garment, fabric, labelling, and delivery instructions for the individual order."),
    ("Tab: Body Measurements", "Raw measurements taken on the customer, in centimetres. Jacket, shirt and trouser blocks."),
    ("Tab: Finished Measurements", "Target finished-garment measurements after ease. Supplied when the customer has an approved fit garment on file."),
    ("Tab: Craft Specification", "The construction selections for this order, given as Specification Codes from the Craft Options Catalogue."),
    ("Specification codes", "Codes such as 'lapel-peak-9' are drawn from the Craft Options Catalogue workbook supplied alongside this file."),
    ("Units", "Centimetres throughout. Customers entering imperial are converted at intake before the order is transmitted."),
    ("Tolerance", "Finished-garment tolerance requested: +/- 0.5 cm on all horizontal measurements, +/- 1.0 cm on lengths."),
]
r = 3
for k, v in lines:
    a = ws.cell(row=r, column=1, value=k)
    a.font, a.fill, a.border = BOLD, SEC_FILL, BOX
    a.alignment = Alignment(vertical="top", indent=1)
    b = ws.cell(row=r, column=2, value=v)
    b.font, b.border = BODY, BOX
    b.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 30
    r += 1

lg = ws.cell(row=r + 1, column=1, value="LEGEND")
lg.font = BOLD
c = ws.cell(row=r + 2, column=1, value="Per-order input")
c.fill, c.font, c.border = INPUT_FILL, BODY, BOX
ws.cell(row=r + 2, column=2, value="Completed by Blessed & Dressed for each individual garment.").font = BODY
c = ws.cell(row=r + 3, column=1, value="Fixed template")
c.fill, c.font, c.border = SEC_FILL, BODY, BOX
ws.cell(row=r + 3, column=2, value="Standing instruction, identical on every order.").font = BODY

# ---- Sheet: Order Header ----
ws = wb2.create_sheet("Order Header")
banner(ws, "ORDER HEADER  —  EXAMPLE ORDER BD-2026-0417", 3)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 52
ws.column_dimensions["C"].width = 60

header_row(ws, 3, ["Field", "Value (example)", "Notes"], [34, 52, 60])
fields = [
    ("Order number", "BD-2026-0417", "Blessed & Dressed order reference. Must appear on the garment bag and packing list."),
    ("Order date", "2026-09-02", "ISO format."),
    ("Customer reference", "CUST-00318", "Anonymised customer key. Pattern is retained under this key for repeat orders."),
    ("Garment type", "Suit (2-Piece)", "Matches a tab in the Craft Options Catalogue."),
    ("Quantity", 1, "Single-garment order. Standard for our model."),
    ("Fit block", "Blessed & Dressed Block B — Classic", "House block. Supplied to you as a graded pattern set."),
    ("Fabric — supplier", "Customer's own / nominated mill", "We expect to nominate fabric per order; confirm whether you accept CMT."),
    ("Fabric — reference", "S120 Plain Black — 3.6 m", "Mill reference and yardage allocated to this order."),
    ("Lining reference", "Bemberg cupro — Jet Black", ""),
    ("Canvas construction", "Full canvas", "See Craft Specification tab for the exact code."),
    ("Brand label", "Blessed & Dressed — woven, black on ivory", "Private label. Artwork supplied on approval."),
    ("Care / composition label", "Required — U.S. FTC compliant", "English, fibre content, country of origin, RN pending."),
    ("Monogram", "D.S. — interior left facing, ivory thread", "Optional per order."),
    ("Target lead time", "21 working days ex-works", "Please confirm achievable lead time."),
    ("Ship to", "Blessed & Dressed, Louisiana, United States", "DDP quotation requested; DAP acceptable as alternate."),
    ("QC requirement", "Finished-garment measurement report per unit", "Photographed on form, front and back, before packing."),
]
r = 4
for k, v, n in fields:
    a = ws.cell(row=r, column=1, value=k)
    a.font, a.fill, a.border = BOLD, SEC_FILL, BOX
    a.alignment = Alignment(vertical="center", indent=1)
    b = ws.cell(row=r, column=2, value=v)
    b.font, b.fill, b.border = BODY, INPUT_FILL, BOX
    b.alignment = Alignment(vertical="center", indent=1)
    c = ws.cell(row=r, column=3, value=n)
    c.font, c.border = SMALL, BOX
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    ws.row_dimensions[r].height = 24
    r += 1

# ---- Measurement sheets ----
EXAMPLE = {
    # jacket / body
    "height": 183.0, "weight": 84.0, "chest": 104.0, "stomach": 92.0, "belly": 94.0,
    "seat": 101.0, "shoulder": 46.5, "back_width": 44.0, "bicep": 33.0, "cuff": 18.0,
    "left_sleeve": 64.5, "right_sleeve": 65.0, "back_length": 76.0, "front_length": 71.5,
    "first_button": 47.0, "neck": 40.0, "left_wrist": 17.5, "right_wrist": 17.5,
    "nape_to_waist": 45.0, "waist": 90.0, "thigh": 60.0, "rise": 27.0, "inseam": 82.0,
    "knee": 44.0, "bottom": 36.0, "length": 106.0, "sleeve": 64.5, "collar": 40.0,
}


def meas_sheet(sheet_name, title, subtitle, blocks):
    s = wb2.create_sheet(sheet_name)
    banner(s, title, 5)
    s.cell(row=2, column=1, value=subtitle).font = SMALL
    s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    header_row(s, 4, ["Block", "Measurement", "Value", "Unit", "How It Is Taken"],
               [22, 30, 14, 10, 92])
    r = 5
    for block_label, rows in blocks:
        for m in rows:
            a = s.cell(row=r, column=1, value=block_label)
            a.font, a.fill, a.border = BOLD, SEC_FILL, BOX
            a.alignment = Alignment(vertical="top", indent=1)
            b = s.cell(row=r, column=2, value=m["label"])
            b.font, b.border = BODY, BOX
            b.alignment = Alignment(vertical="top", indent=1)
            v = s.cell(row=r, column=3, value=EXAMPLE.get(m["key"], ""))
            v.font, v.fill, v.border = BODY, INPUT_FILL, BOX
            v.number_format = "0.0"
            v.alignment = Alignment(horizontal="center", vertical="top")
            unit = "kg" if m["key"] == "weight" else "cm"
            k = s.cell(row=r, column=4, value=unit)
            k.font, k.border = BODY, BOX
            k.alignment = Alignment(horizontal="center", vertical="top")
            h = s.cell(row=r, column=5, value=m["how"])
            h.font, h.border = SMALL, BOX
            h.alignment = Alignment(vertical="top", wrap_text=True, indent=1)
            s.row_dimensions[r].height = 28
            r += 1
    return s, r


meas_sheet(
    "Body Measurements",
    "BODY MEASUREMENTS  —  TAKEN ON THE CUSTOMER  (cm)",
    "Raw body measurements, centimetres except weight. Ease is applied by the pattern, not by the measurer. Example values are for order BD-2026-0417.",
    [("Jacket / Coat", meas["BODY_JACKET"]),
     ("Shirt", meas["BODY_SHIRT"]),
     ("Trousers", meas["BODY_TROUSERS"])],
)

s, r = meas_sheet(
    "Finished Measurements",
    "FINISHED GARMENT MEASUREMENTS  —  TARGET AFTER EASE  (cm)",
    "Supplied when the customer has an approved fit garment on file. These are the numbers we ask to be verified on the finished piece before packing.",
    [("Jacket / Coat", meas["FINISHED_JACKET"]),
     ("Shirt", meas["FINISHED_SHIRT"]),
     ("Trousers", meas["FINISHED_TROUSERS"])],
)
n = s.cell(row=r + 1, column=1,
           value="Requested tolerance: +/- 0.5 cm on horizontal measurements, +/- 1.0 cm on lengths. Please confirm your standard tolerance if it differs.")
n.font = SMALL
s.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)

# ---- Sheet: Craft Specification ----
suit = next(p for p in products if p["name"] == "Suit (2-Piece)")
defaults = [x for x in suit["rows"] if x["isDefault"] == "Yes"]

s = wb2.create_sheet("Craft Specification")
banner(s, "CRAFT SPECIFICATION  —  EXAMPLE ORDER BD-2026-0417  (SUIT, 2-PIECE)", 5)
s.cell(row=2, column=1,
       value="One line per construction decision. Codes are drawn from the Craft Options Catalogue. A real order transmits every line below.").font = SMALL
s.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
header_row(s, 4, ["Section", "Option Category", "Selection", "Specification Code", "Tech Illustration On File"],
           [28, 30, 34, 32, 24])
r = 5
for row in defaults:
    vals = [row["section"], row["field"], row["name"], row["code"], row["illustration"]]
    for i, v in enumerate(vals, start=1):
        c = s.cell(row=r, column=i, value=v)
        c.font, c.border = BODY, BOX
        c.alignment = Alignment(vertical="center", indent=1)
        if i <= 2:
            c.fill = SEC_FILL
            c.font = BOLD
        if i in (3, 4):
            c.fill = INPUT_FILL
        if i == 5:
            c.alignment = Alignment(horizontal="center")
    s.row_dimensions[r].height = 20
    r += 1
s.auto_filter.ref = f"A4:E{r-1}"
note = s.cell(row=r + 1, column=1,
              value=f"{len(defaults)} specification lines on this example order. The full option set for this garment is {len(suit['rows'])} options across {suit['sections']} groups — see the Craft Options Catalogue.")
note.font = SMALL
s.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=5)

wb2.save(os.path.join(OUTDIR, "Blessed-and-Dressed_Sample-Order-Form.xlsx"))
print("order form workbook written; spec lines:", len(defaults))
