#!/usr/bin/env python3
"""Extract tech-pack illustrations + build catalog.json from the supplier workbook.

Usage:
    python extract_xlsx_assets.py "<path to shirt design.xlsx>" [--out shirt-assets]

Reads the "Design Options Map" sheet (canonical option table) and the "Design Details"
sheet (category header rows with one embedded illustration per option row), writes every
illustration to <out>/illustrations/<category_key>/<slug>.<ext>, and creates/merges
<out>/catalog.json. Safe to re-run: existing statuses, photos, and user-edited `kind`
values are preserved. Prints per-category counts and loudly flags any mismatch between
expected options and found images — investigate mismatches, never guess.

Requires: openpyxl  (pip install openpyxl)
"""
import argparse
import json
import re
import sys
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl --break-system-packages")

MAP_SHEET = "Design Options Map"
DETAILS_SHEET = "Design Details"

# Categories whose images are color chips / hardware pickers, not construction blueprints.
# These keep their extracted illustration as the display asset (no AI generation).
SWATCH_KEYS = {
    "button_on_collar_stand", "placket_button", "collar_linings", "cuff_lining",
    "placket_splicing_material", "contrast_fabric",
}


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def slugify(label):
    s = re.sub(r"[^a-z0-9]+", "-", str(label).lower()).strip("-")
    return s or "option"


def sniff_ext(data):
    if data[:2] == b"\xff\xd8":
        return ".jpg"
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return ".png"
    if data[:4] in (b"GIF8",):
        return ".gif"
    return ".img"


def classify_kind(category_key):
    if "color" in category_key or category_key in SWATCH_KEYS:
        return "swatch"
    return "structural"


def parse_measurements(label):
    """Pull cm values out of an option label like 'Fashion point in 5.8cm'."""
    return [float(m) for m in re.findall(r"(\d+(?:\.\d+)?)\s*cm", str(label), re.I)]


def read_map(wb):
    """Ordered option rows from Design Options Map."""
    ws = wb[MAP_SHEET]
    rows, categories = [], []  # categories in first-appearance order
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or r[0] is None or not str(r[0]).strip():
            continue
        key = str(r[0]).strip()
        label = str(r[1]).strip() if r[1] else key
        required = str(r[2]).strip().upper() == "Y" if r[2] else False
        try:
            option_id = int(r[3])
        except (TypeError, ValueError):
            continue
        option_label = str(r[4]).strip() if r[4] else f"Option {option_id}"
        image_path = str(r[5]).strip() if len(r) > 5 and r[5] else ""
        if key not in [c[0] for c in categories]:
            categories.append((key, label))
        rows.append({
            "category_key": key, "category_label": label, "required": required,
            "option_id": option_id, "option_label": option_label,
            "map_image_path": image_path,
        })
    return rows, categories


def read_details(wb):
    """Header rows (text) + images (anchor rows) from Design Details."""
    ws = wb[DETAILS_SHEET]
    headers = []  # (row, text)
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and str(c.value).strip():
                headers.append((c.row, str(c.value).strip()))
    images = []  # (row, image_object)
    for img in ws._images:
        images.append((img.anchor._from.row + 1, img))
    images.sort(key=lambda t: t[0])
    headers.sort(key=lambda t: t[0])
    return headers, images


def pair_headers_to_categories(headers, categories):
    """Sheet headers appear in the same order as map categories. Verify by fuzzy label
    match; the first header is known-truncated ('leeve head'), so order is primary and
    label similarity is the alarm bell."""
    pairs, warnings = [], []
    n = min(len(headers), len(categories))
    if len(headers) != len(categories):
        warnings.append(f"header/category count differs: {len(headers)} headers vs "
                        f"{len(categories)} categories — pairing first {n} in order")
    for i in range(n):
        (row, text), (key, label) = headers[i], categories[i]
        a, b = norm(text.rstrip("*")), norm(label)
        sim = SequenceMatcher(None, a, b).ratio()
        if not (a in b or b in a or sim > 0.55):
            warnings.append(f"header '{text}' (row {row}) paired with category "
                            f"'{label}' but labels look different (sim={sim:.2f}) — verify")
        pairs.append({"row": row, "category_key": key})
    return pairs, warnings


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", default="shirt-assets")
    args = ap.parse_args()

    out = Path(args.out)
    (out / "illustrations").mkdir(parents=True, exist_ok=True)
    (out / "reports").mkdir(parents=True, exist_ok=True)
    catalog_path = out / "catalog.json"

    print(f"Loading {args.xlsx} ...")
    wb = openpyxl.load_workbook(args.xlsx)
    map_rows, categories = read_map(wb)
    headers, images = read_details(wb)
    pairs, warnings = pair_headers_to_categories(headers, categories)

    # assign images to categories by header row ranges
    by_cat = {p["category_key"]: [] for p in pairs}
    bounds = [(p["row"], (pairs[i + 1]["row"] if i + 1 < len(pairs) else 10**9),
               p["category_key"]) for i, p in enumerate(pairs)]
    for row, img in images:
        assigned = None
        for start, end, key in bounds:
            if start <= row < end:
                assigned = key
                break
        if assigned:
            by_cat[assigned].append((row, img))
        else:
            warnings.append(f"image at row {row} sits above the first header — skipped")

    # existing catalog (merge)
    existing = {}
    if catalog_path.exists():
        old = json.loads(catalog_path.read_text(encoding="utf-8"))
        existing = {o["product_code"]: o for o in old.get("options", [])}

    options_out, summary = [], []
    for key, label in categories:
        opts = [r for r in map_rows if r["category_key"] == key]
        imgs = by_cat.get(key, [])
        n = min(len(opts), len(imgs))
        if len(opts) != len(imgs):
            warnings.append(f"[{key}] map has {len(opts)} options but sheet has "
                            f"{len(imgs)} images — wrote first {n} pairs, REVIEW THIS")
        wrote = 0
        for i, opt in enumerate(opts):
            code = f"{key}-{opt['option_id']:02d}"
            slug = (Path(opt["map_image_path"]).stem if opt["map_image_path"]
                    else slugify(opt["option_label"]))
            entry = {
                "product_code": code, "category_key": key, "category_label": label,
                "required": opt["required"], "option_id": opt["option_id"],
                "option_label": opt["option_label"], "slug": slug,
                "kind": classify_kind(key),
                "measurements_cm": parse_measurements(opt["option_label"]),
                "illustration": None, "photo": None, "photo_annotated": None,
                "spec": None, "qa": None, "status": "pending", "attempts": 0,
                "updated": str(date.today()),
            }
            if i < n:
                data = imgs[i][1]._data()
                ext = sniff_ext(data)
                rel = f"illustrations/{key}/{slug}{ext}"
                p = out / rel
                p.parent.mkdir(parents=True, exist_ok=True)
                p.write_bytes(data)
                entry["illustration"] = rel
                wrote += 1
            else:
                entry["status"] = "missing-illustration"
            # merge: keep human/pipeline progress from previous runs
            if code in existing:
                keep = existing[code]
                for f in ("kind", "photo", "photo_annotated", "spec", "qa",
                          "status", "attempts"):
                    if keep.get(f) not in (None, 0, "pending", "skipped-swatch"):
                        entry[f] = keep[f]
            if entry["kind"] == "swatch" and entry["status"] == "pending":
                entry["status"] = "skipped-swatch"
            options_out.append(entry)
        summary.append((key, label, len(opts), len(imgs), wrote))

    catalog = {
        "generated": datetime.now().isoformat(timespec="seconds"),
        "source": Path(args.xlsx).name,
        "options": options_out,
    }
    catalog_path.write_text(json.dumps(catalog, indent=2, ensure_ascii=False),
                            encoding="utf-8")

    print(f"\n{'category':<42}{'map':>5}{'imgs':>6}{'wrote':>7}")
    for key, label, o, im, w in summary:
        flag = "  <-- MISMATCH" if o != im else ""
        print(f"{key:<42}{o:>5}{im:>6}{w:>7}{flag}")
    total_o = sum(s[2] for s in summary)
    total_w = sum(s[4] for s in summary)
    kinds = {}
    for e in options_out:
        kinds[e["kind"]] = kinds.get(e["kind"], 0) + 1
    print(f"\nTotal: {total_o} options, {total_w} illustrations written")
    print(f"Kinds: {kinds}  (flip any entry's kind in catalog.json if I guessed wrong)")
    print(f"Catalog: {catalog_path}")
    if warnings:
        print("\nWARNINGS:")
        for w in warnings:
            print(f"  ! {w}")
    note = ("Note: 'Monogram' and 'Body Adjustment' sheets also contain images/maps; "
            "extract them later with the same pattern if needed.")
    print(f"\n{note}")


if __name__ == "__main__":
    main()
